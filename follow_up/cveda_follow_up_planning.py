#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# Copyright (c) 2017 CEA
#
# This software is governed by the CeCILL license under French law and
# abiding by the rules of distribution of free software. You can use,
# modify and/ or redistribute the software under the terms of the CeCILL
# license as circulated by CEA, CNRS and INRIA at the following URL
# "http://www.cecill.info".
#
# As a counterpart to the access to the source code and rights to copy,
# modify and redistribute granted by the license, users are provided only
# with a limited warranty and the software's author, the holder of the
# economic rights, and the successive licensors have only limited
# liability.
#
# In this respect, the user's attention is drawn to the risks associated
# with loading, using, modifying and/or developing or reproducing the
# software by the user in light of its specific status of free software,
# that may mean that it is complicated to manipulate, and that also
# therefore means that it is reserved for developers and experienced
# professionals having in-depth computer knowledge. Users are therefore
# encouraged to load and test the software's suitability as regards their
# requirements in conditions enabling the security of their systems and/or
# data to be ensured and, more generally, to use and operate it in the
# same conditions as regards security.
#
# The fact that you are presently reading this means that you have had
# knowledge of the CeCILL license and that you accept its terms.


from openpyxl import load_workbook, Workbook
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from random import shuffle
import logging
logger = logging.getLogger()


# reference file for date of birth, sex, date of baseline assessment
_EXCEL_REFERENCE_PATH = '/cveda/databank/framework/meta_data/Recruitment sheet for Follow Up.xlsx'

_EXCEL_REFERENCE_HEADER = [
    'Sl no',
    'PSC1 Code',
    'DOB',
    'Age',
    'Gender',
    'Age Band',
    'Assessment date'
]

_AGE_BAND = {
    'C1': 1,
    'C2': 2,
    'C3': 3,
}


def _check_excel_reference_header(header):
    return list(header) == _EXCEL_REFERENCE_HEADER


def _remove_line_breaks(s):
    for c in ('\n','\r'):
        if c in s:
            s = s.replace(c, '')
    return s


def _age_band(age_in_years):
    # As documented here: https://github.com/cveda/cveda_databank/wiki
    #   C1:  6-11
    #   C2: 12-17
    #   C3: 18-23
    # The few subjects < 6 and > 23 end up in C1 and C3 respectively.
    if age_in_years < 12:
        return 1
    elif age_in_years < 18:
        return 2
    else:
        return 3


def _season(d):
    # As documented here: https://en.wikipedia.org/wiki/Climate_of_India
    #   The nation has four seasons:
    #   winter (December, January and February),
    #   summer (March, April and May),
    #   a monsoon rainy season (June to September),
    #   and a post-monsoon period (October to November).
    if d.month in {12, 1, 2}:
        return 1
    elif d.month in {3, 4, 5}:
        return 2
    elif d.month in {6, 7, 8, 9}:
        return 3
    else:  # in {10, 11}
        return 4


def read_excel_reference(path):
    workbook = load_workbook(path)
    data = {}
    for worksheet in workbook:
        # The Excel file contains a sheet per acquisition centre.
        center = worksheet.title
        if center == 'MYSORE':
            center = 'MYSURU'
        elif center == 'PGIMER':
            center = 'CHANDIGARH'
        elif center == 'KOLKATA':
            continue  # still empty

        # Process each sheet.
        header = None
        for row in worksheet.iter_rows():
            if not header:
                # The 1st row is a header with column names.
                header = [_remove_line_breaks(cell.value.strip())
                          for cell in row]
                if not _check_excel_reference_header(header):
                    print('ERROR: unexpected header: ' + ', '.join(header))
            else:
                # PSC1 Code
                psc1 = row[1]
                if psc1.data_type == 's' or psc1.data_type == 'n':
                    if psc1.data_type == 'n':
                        psc1 = str(psc1.value)
                    else:
                        psc1 = psc1.value
                    if len(psc1) != 12:
                        logger.warning('%s: invalid "PSC1 Code"' % psc1)
                        psc1 = None
                else:
                    logger.warning('%s: incorrect "PSC1 Code" type: %s'
                                   % (center, psc1.data_type))
                    psc1 = None

                # DOB
                dob = row[2]
                if dob.data_type == 'n' and dob.is_date:
                    dob = dob.value
                    if type(dob) == datetime:
                        dob = dob.date()
                elif dob.data_type == 's' and dob.value == 'INITIATED':
                    dob = None
                else:
                    logger.warning('%s: incorrect "DOB" type: %s'
                                   % (psc1, dob.data_type))
                    dob = None

                # Assessment date
                assessment = row[6]
                if assessment.value == None:
                    assessment = None
                elif assessment.data_type == 'n' and assessment.is_date:
                    assessment = assessment.value
                    if type(assessment) == datetime:
                        assessment = assessment.date()
                    else:
                        logger.warning('%s: unexpected "Assessment date" type: %s'
                                       % (psc1, type(assessment)))
                        assessment = None
                else:
                    logger.warning('%s: incorrect "Assessment date": %s'
                                   % (psc1, assessment.value))
                    assessment = None

                # Define the "season" of assessment.
                if assessment:
                    season = _season(assessment)
                else:
                    season = None

                # Gender/Sex
                sex = row[4]
                if sex.data_type == 's':
                    sex = sex.value
                    if sex not in {'F', 'M'}:
                        logger.warning('%s: invalid "Sex": %s' % (psc1, sex))
                        sex = None
                else:
                    if sex.value != None:
                        logger.warning('%s: incorrect "Sex" type: %s'
                                       % (psc1, sex.data_type))
                    sex = None

                # Calculate theoretical age in years at baseline.
                if dob and assessment:
                    calculated_age = relativedelta(assessment, dob).years
                else:
                    calculated_age = None

                # Age
                age = row[3]
                if age.data_type == 'n':
                    age = age.value
                    # double check against calculated age
                    if calculated_age and calculated_age != age and abs(calculated_age - age) > 2:
                        logger.error('%s: incorrect "Age": %d (%d)'
                                     % (psc1, age, calculated_age))
                elif age.data_type == 'f':
                    pass  # typically '=DATEDIF(date_of_birth,assessment_date,"Y")'
                    age = calculated_age
                else:
                    if age.value != None:
                        logger.warning('%s: incorrect "Age" type: %s'
                                       % (psc1, age.data_type))
                    age = None

                # Calculate age band at baseline.
                if age:
                    calculated_band = _age_band(age)
                else:
                    calculated_band = None

                # Age Band
                band = row[5]
                if band.data_type == 's':
                    band = band.value
                    if band in _AGE_BAND:
                        band = _AGE_BAND[band]
                        # double check against calculated age band
                        if age and band != _age_band(age):
                            logger.error('%s: incorrect "Age Band": C%d (C%d)'
                                           % (psc1, band, _age_band(age)))
                    else:
                        logger.warning('%s: invalid "Age Band": %s' % (psc1, band))
                        band = None
                else:
                    if band.value != None:
                        logger.warning('%s: incorrect "Age Band" type: %s'
                                       % (psc1, band.data_type))
                    band = None

                # Final consistency check: cannot have Age Band without Age!
                if band and not age:
                    logger.error('%s: "Age Band" without "Age": C%d'
                                 % (psc1, band))

                if calculated_band and sex and season:
                    data.setdefault(center, {}) \
                        .setdefault(calculated_band, {}) \
                        .setdefault(sex, {}) \
                        .setdefault(season, {})[psc1] =  (dob, sex, assessment)

    return data


def randomize(data):
    follow_ups = ({}, {})  # follow-up 1 and 2

    for center, bands in data.items():
        for band, sexes in bands.items():
            for sex, seasons in sexes.items():
                for season, subjects in seasons.items():
                    psc1s = list(subjects.keys())
                    # Randomly split each gpoup in half.
                    half_len = len(psc1s) // 2
                    shuffle(psc1s)
                    partitions = [psc1s[:half_len], psc1s[half_len:]]
                    shuffle(partitions)
                    for i in (0, 1):  # follow-up 1 and 2
                        follow_ups[i].setdefault(center, {}) \
                            .update([(psc1, subjects[psc1])
                                     for psc1 in partitions[i]])

    return follow_ups


def main():
    # Read reference Excel file into 192 groups of subjects:
    #   8 centres × 3 age bands × 2 × 4 seasons
    data = read_excel_reference(_EXCEL_REFERENCE_PATH)

    # Randomly split each group into follow-up 1 / 2.
    # Return a dictionary of PSC1 codes for follow-up 1 / 2.
    follow_ups = randomize(data)

    # Prepare nice output.
    today = datetime.now()
    for i in (0, 1):  # follow-up 1 and 2
        workbook = Workbook()
        workbook.remove_sheet(workbook.active)
        for center, subjects in follow_ups[i].items():
            worksheet = workbook.create_sheet(title=center)
            # sort subjects by assessment date
            for psc1, (dob, sex, assessment) in sorted(subjects.items(),
                                                       key=lambda x: x[1][2]):
                # theoretical follow-up assessment date is 1 / 2 years later
                years = i + 1
                # calculate new values at follow-up
                new_assessment = assessment + relativedelta(years=years)
                age = relativedelta(new_assessment, dob).years
                band = _age_band(age)
                calendar_month = assessment.strftime('%B %Y')
                worksheet.append([calendar_month, new_assessment,
                                  psc1, dob, sex,
                                  age, 'C{}'.format(band)])
        workbook.save(filename='follow_up_{}_{}.xlsx'
                               .format(i + 1, today.strftime('%Y-%m-%d')))


if __name__ == "__main__":
    main()
