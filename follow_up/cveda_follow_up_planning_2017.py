#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# Copyright (c) 2017 CEA
#
# This software is governed by the CeCILL license under French law and
# abiding by the rules of distribution of free software. You can use,
# modify and/ or redistribute the software under the terms of the CeCILL
# license as circulated by CEA, CNRS and INRIA at the following URL
# "http://www.cecill.info".
#
# As a counterpart to the access to the source code and rights to copy,
# modify and redistribute granted by the license, users are provided only
# with a limited warranty and the software's author, the holder of the
# economic rights, and the successive licensors have only limited
# liability.
#
# In this respect, the user's attention is drawn to the risks associated
# with loading, using, modifying and/or developing or reproducing the
# software by the user in light of its specific status of free software,
# that may mean that it is complicated to manipulate, and that also
# therefore means that it is reserved for developers and experienced
# professionals having in-depth computer knowledge. Users are therefore
# encouraged to load and test the software's suitability as regards their
# requirements in conditions enabling the security of their systems and/or
# data to be ensured and, more generally, to use and operate it in the
# same conditions as regards security.
#
# The fact that you are presently reading this means that you have had
# knowledge of the CeCILL license and that you accept its terms.


from openpyxl import load_workbook, Workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta
from random import shuffle
import logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger()


# reference file for date of birth, sex, date of baseline assessment
_EXCEL_REFERENCE_PATH = '/cveda/databank/framework/meta_data/follow_up/randomization_2017/recruitment_file_2017-11-03.xlsx'

_EXCEL_REFERENCE_HEADER = [
    'Sl no',  # A
    'PSC1 Code',  # B
    'DOB',  # C
    'Age',  # D
    'Sex',  # E
    'Age Band',  # F
    'Assessment date',  # G
    'MRI',  # H
    'Correct age',  # I
    'Age Difference',  # J
    'Remarks',  # K
    "PSC1 code matching from Dimitri's list",  # L
    'Age discrepancy present',  # M
    'Age band changes',  # N
    "PSC1 codes not matching in Dimitri's list",  # O
    'Comment',  # P
    'Action',  # Q
]

_AGE_BAND = {
    'C1': 1,
    'C2': 2,
    'C3': 3,
}

_CENTERS = {  # select centers and discard spurious sheets in Excel file
    'CHANDIGARH',
    'MANIPUR',
    'MYSORE',
    'NIMHANS',
    'SJRI',
    # 'KOLKATA',
}


def _check_excel_reference_header(header):
    return list(header) == _EXCEL_REFERENCE_HEADER


def _age_band(age_in_years):
    # As documented here: https://github.com/cveda/cveda_databank/wiki
    #   C1:  6-11
    #   C2: 12-17
    #   C3: 18-23
    # The few subjects < 6 and > 23 end up in C1 and C3 respectively.
    if age_in_years < 12:
        return 1
    elif age_in_years < 18:
        return 2
    else:
        return 3


def _time_block(d):
    # Ensure that we don't get all the people in the first follow up
    # coming from the same part of the year.
    if d.month in {11, 12, 1}:
        return 1
    elif d.month in {2, 3, 4}:
        return 2
    elif d.month in {5, 6, 7}:
        return 3
    else:  # if d.month in {8, 9, 10}
        return 4


def process_worksheet(worksheet):
    data = {}

    # spurious sheets *must* have been filtered out at this point
    center = worksheet.title

    seen = set()  # detect duplicates
    discard = set()  # discard PSC1 entries yet to be fixed

    header = None
    for row in worksheet.iter_rows():
        if not header:
            # The 1st row is a header with column names.
            header = [cell.value.strip() for cell in row]
            if not _check_excel_reference_header(header):
                logger.error('%s: unexpected header: %s'
                             % (center, ', '.join(header)))
        else:
            # PSC1 Code
            psc1 = row[1]  # B
            if psc1.data_type == 's' or psc1.data_type == 'n':
                if psc1.data_type == 'n':
                    psc1 = str(psc1.value)
                else:
                    psc1 = psc1.value
                if psc1 in seen:
                    logger.error('%s: duplicate "PSC1 Code"' % psc1)
                else:
                    logger.info('%s: first occurrence of this "PSC1 Code"' % psc1)
                    seen.add(psc1)
                if len(psc1) != 12:
                    logger.error('%s: invalid "PSC1 Code"' % psc1)
                    psc1 = None
            else:
                logger.error('%s: incorrect "PSC1 Code" type "%s": %s'
                             % (center, psc1.data_type, psc1.value))
                psc1 = None

            # DOB
            dob = row[2]  # C
            if dob.data_type == 'n' and dob.is_date:
                dob = dob.value
                if type(dob) == datetime:
                    dob = dob.date()
            elif dob.data_type == 's' and dob.value == 'INITIATED':
                dob = None
            else:
                logger.error('%s: incorrect "DOB" type "%s": %s'
                             % (psc1, dob.data_type, dob.value))
                dob = None

            # Assessment date
            assessment = row[6]  # G
            if assessment.value is None:
                assessment = None
            elif assessment.data_type == 'n' and assessment.is_date:
                assessment = assessment.value
                if type(assessment) == datetime:
                    assessment = assessment.date()
                else:
                    logger.error('%s: unexpected "Assessment date" type: %s'
                                 % (psc1, type(assessment)))
                    assessment = None
            else:
                logger.warning('%s: incorrect "Assessment date" "%s": %s'
                               % (psc1, assessment.data_type, assessment.value))
                assessment = None

            # Define the time-block of assessment.
            if assessment:
                time_block = _time_block(assessment)
            else:
                time_block = None

            # Sex
            sex = row[4]  # E
            if sex.data_type == 's':
                sex = sex.value
                if sex not in {'F', 'M'}:
                    logger.error('%s: invalid "Sex": %s' % (psc1, sex))
                    sex = None
            else:
                if sex.value is not None:
                    logger.error('%s: incorrect "Sex" type "%s": %s'
                                 % (psc1, sex.data_type, sex.value))
                sex = None

            # Calculate theoretical age in years at baseline.
            if dob and assessment:
                calculated_age = relativedelta(assessment, dob).years
            else:
                calculated_age = None

            # Age
            age = row[3]  # D
            if age.data_type == 'n':
                age = age.value
                # double check against calculated age
                if calculated_age and calculated_age != age:
                    logger.error('%s: incorrect "Age": %d (%d)'
                                 % (psc1, age, calculated_age))
            elif age.data_type == 'f':
                # typically '=DATEDIF(date_of_birth, assessment_date, "Y")'
                age = None
            elif age.value is not None:
                logger.error('%s: incorrect "Age" type "%s": %s'
                             % (psc1, age.data_type, age.value))
                age = None

            #  Correct age
            correct_age = row[8]  # I
            if correct_age.data_type == 'n':
                correct_age = correct_age.value
                # double check against calculated age
                if calculated_age and calculated_age != correct_age:
                    logger.error('%s: incorrect "Correct age": %d (%d)'
                                 % (psc1, correct_age, calculated_age))
            elif correct_age.data_type == 'f':
                # typically '=INT((assessment_date - date_of_birth) / 365)'
                correct_age = None
            elif correct_age.value is not None:
                logger.error('%s: incorrect "Correct age" type "%s": %s'
                             % (psc1, correct_age.data_type, correct_age.value))
                correct_age = None

            # Calculate age band at baseline.
            if calculated_age:
                calculated_band = _age_band(calculated_age)
                logger.info('%s: age band from calculated age: %d -> C%d'
                               % (psc1, calculated_age, calculated_band))
            elif correct_age:
                calculated_band = _age_band(correct_age)
                logger.warning('%s: age band from "Correct age": %d -> C%d'
                               % (psc1, correct_age, calculated_band))
            elif age:
                calculated_band = _age_band(age)
                logger.warning('%s: age band from "Age": %d -> C%d'
                               % (psc1, age, calculated_band))
            else:
                logger.error('%s: cannot determine age band' % psc1)
                calculated_band = None

            # Age Band
            band = row[5]  # F
            if band.data_type == 's':
                band = band.value
                if band in _AGE_BAND:
                    band = _AGE_BAND[band]
                    # double check against calculated age band
                    if calculated_band and band != calculated_band:
                        logger.error('%s: incorrect "Age Band": C%d (C%d)'
                                       % (psc1, band, calculated_band))
                else:
                    logger.error('%s: invalid "Age Band": %s' % (psc1, band))
                    band = None
            else:
                if band.value is not None:
                    logger.error('%s: incorrect "Age Band" type "%s": %s'
                                 % (psc1, band.data_type, band.value))
                band = None

            # PSC1 codes not matching in Dimitri's list
            bogus = row[14]  # O
            if bogus.data_type == 's' or bogus.data_type == 'n':
                if bogus.data_type == 'n':
                    bogus = str(bogus.value)
                else:
                    bogus = bogus.value
                if bogus and len(bogus) != 12:
                    logger.error('%s: invalid bogus "PSC1 Code"' % bogus)
                    bogus = None
            else:
                logger.error('%s: incorrect bogus "PSC1 Code" type "%s": %s'
                             % (center, bogus.data_type, bogus.value))
                bogus = None
            if bogus:
                logger.info('%s: bogus "PSC1 Code" will be discarded' % bogus)
                discard.add(bogus)

            if calculated_band and sex and time_block:
                data.setdefault(calculated_band, {}) \
                    .setdefault(sex, {}) \
                    .setdefault(time_block, {})[psc1] = (dob, sex, assessment)

    for band, sexes in data.items():
        for sex, time_blocks in sexes.items():
            for time_block, subjects in time_blocks.items():
                for key in discard.intersection(subjects.keys()):
                    logger.warning('%s: discard!' % key)
                    del subjects[key]

    return center, data


def read_excel_reference(path):
    workbook = load_workbook(path)

    return {center: data
            for (center, data) in [process_worksheet(worksheet)
                                   for worksheet in workbook
                                   if worksheet.title in _CENTERS]}


def randomize(data):
    follow_ups = ({}, {})  # follow-up 1 and follow-up 2

    for center, bands in data.items():
        for sexes in bands.values():
            for time_blocks in sexes.values():
                for subjects in time_blocks.values():
                    psc1s = list(subjects.keys())
                    # Randomly split each gpoup in half.
                    half_len = len(psc1s) // 2
                    shuffle(psc1s)
                    partitions = [psc1s[:half_len], psc1s[half_len:]]
                    shuffle(partitions)
                    for i in (0, 1):  # follow-up 1 and follow-up 2
                        follow_ups[i].setdefault(center, {}) \
                            .update([(psc1, subjects[psc1])
                                     for psc1 in partitions[i]])

    return follow_ups


def main():
    # Read reference Excel file into 192 groups of subjects:
    #   8 centres × 3 age bands × 2 × 4 time-blocks
    data = read_excel_reference(_EXCEL_REFERENCE_PATH)

    # Randomly split each group into follow-up 1 / follow-up 2.
    # Return a dictionary of PSC1 codes for follow-up 1 / follow-up 2.
    follow_ups = randomize(data)

    # Prepare nice output.
    today = datetime.now()
    for i in (0, 1):  # follow-up 1 and follow-up 2
        workbook = Workbook()
        workbook.remove_sheet(workbook.active)
        for center, subjects in follow_ups[i].items():
            worksheet = workbook.create_sheet(title=center)
            worksheet.append(['Follow-up assessment month', 'Baseline assessment date',
                              'PSC1', 'Date of Birth', 'Sex',
                              'Planned age', 'Planned age band'])
            # sort subjects by assessment date
            for psc1, (dob, sex, assessment) in sorted(subjects.items(),
                                                       key=lambda x: x[1][2]):
                # theoretical follow-up assessment date is 1 or 2 years later
                years = i + 1
                # calculate new values at follow-up
                new_assessment = assessment + relativedelta(years=years)
                age = relativedelta(new_assessment, dob).years
                band = _age_band(age)
                calendar_month = new_assessment.strftime('%B %Y')
                worksheet.append([calendar_month, assessment,
                                  psc1, dob, sex,
                                  age, 'C{}'.format(band)])
        workbook.save(filename='follow_up_{}_{}.xlsx'
                               .format(i + 1, today.strftime('%Y-%m-%d')))


if __name__ == "__main__":
    main()
